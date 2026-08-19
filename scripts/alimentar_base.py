"""
Script para Alimentação e Sincronização do Catálogo de Painéis
Lê a aba 'BD' da Planilha Modelo Excel (ou catalogo_paineis.json como fallback)
e atualiza a base SQLite (SQLAlchemy).

Colunas lidas da aba BD:
- Coluna C: Código do Produto (PK Inteira)
- Coluna D: Produto Especificado / Descrição Completa
- Coluna F: Medida Base sem detalhes (ex: 1000 X 11 X 260)
- Coluna G: Peso Unitário (kg)
"""

import os
import sys
import json
import re
import openpyxl

# Adicionar pasta raiz ao sys.path
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, BASE_DIR)

from backend.database import SessionLocal, engine, Base
from backend import models

TEMPLATE_EXCEL_PATH = r"Y:\03 - PAINEL CORRUGADO\Ficha_apontamento_modelo_painel.xlsx"
JSON_CATALOG_PATH = os.path.join(os.path.dirname(__file__), "catalogo_paineis.json")


def extract_products_from_excel(excel_path: str = TEMPLATE_EXCEL_PATH):
    """Extrai os produtos da aba BD do arquivo Excel da planilha modelo."""
    if not os.path.exists(excel_path):
        print(f"[Aviso] Arquivo Excel não encontrado em: {excel_path}")
        return []

    print(f"Lendo aba 'BD' do Excel: {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "BD" not in wb.sheetnames:
        print("[Erro] Aba 'BD' não encontrada na planilha!")
        return []

    ws = wb["BD"]
    rows = list(ws.iter_rows(values_only=True))[1:]  # Pula o cabeçalho
    products = []
    seen_codes = set()

    for r in rows:
        raw_code = r[2]      # Coluna C
        raw_desc = r[3]      # Coluna D
        raw_base = r[5]      # Coluna F
        raw_weight = r[6]    # Coluna G

        if raw_code is None or str(raw_code).strip() in ("", "#N/A", "0", "None"):
            continue

        try:
            code_int = int(float(str(raw_code).replace(",", ".")))
        except (ValueError, TypeError):
            continue

        if code_int <= 0 or code_int in seen_codes:
            continue

        seen_codes.add(code_int)
        desc = str(raw_desc).strip() if raw_desc is not None else ""
        if not desc or desc in ("#N/A", "None"):
            continue

        # Tratamento da Medida Base (Coluna F)
        base_dim = str(raw_base).strip() if raw_base is not None else ""
        if "#VALUE" in base_dim or "#N/A" in base_dim or not base_dim or base_dim == "None":
            # Tenta extrair da descrição D (ex: 'P. CORRUGADO 900 X 11 X 260mm' -> '900 X 11 X 260')
            m = re.search(r"(\d+\s*[xX]\s*\d+\s*[xX]\s*\d+)", desc)
            if m:
                base_dim = m.group(1).upper().replace("X", "X").strip()
            else:
                base_dim = desc.replace("P. CORRUGADO", "").replace("P.CORRUGADO", "").strip()

        # Tratamento do Peso (Coluna G)
        try:
            weight = float(str(raw_weight).replace(",", ".")) if raw_weight is not None else 0.0
            if weight < 0:
                weight = 0.0
        except (ValueError, TypeError):
            weight = 0.0

        products.append({
            "code": code_int,
            "name": desc,
            "specification": desc,
            "dimensions": base_dim,
            "unit_weight_kg": round(weight, 2),
            "nominal_capacity_per_hour": 0.0
        })

    print(f"-> Extraídos {len(products)} produtos válidos da aba BD.")
    return products


def carregar_dados():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    print("=" * 65)
    print("  ALIMENTAÇÃO E SINCRONIZAÇÃO DA BASE DE DADOS - PAINÉIS")
    print("=" * 65)

    try:
        # 1. Garantir Máquinas Padrão
        maquinas_padrao = [
            {"name": "Dobra 1", "code": "DOB-01", "sector": "Painéis", "has_production_control": True},
            {"name": "Dobra 2", "code": "DOB-02", "sector": "Painéis", "has_production_control": True},
            {"name": "Solda Lateral 1", "code": "SOL-LAT-01", "sector": "Painéis", "has_production_control": True},
            {"name": "Solda Lateral 2", "code": "SOL-LAT-02", "sector": "Painéis", "has_production_control": True},
            {"name": "Solda Ponto 1", "code": "SOL-PTO-01", "sector": "Painéis", "has_production_control": True},
            {"name": "Solda Ponto 2", "code": "SOL-PTO-02", "sector": "Painéis", "has_production_control": True},
            {"name": "Revisão", "code": "REV-01", "sector": "Painéis", "has_production_control": False},
            {"name": "Solda Manual", "code": "SOL-MAN-01", "sector": "Painéis", "has_production_control": False},
        ]

        print("\n[1/2] Verificando Máquinas...")
        for m_data in maquinas_padrao:
            maq = db.query(models.Machine).filter(models.Machine.name == m_data["name"]).first()
            if not maq:
                maq = models.Machine(**m_data)
                db.add(maq)
                print(f"  + Cadastrada máquina: {m_data['name']}")
            else:
                maq.code = m_data["code"]
                maq.has_production_control = m_data["has_production_control"]
        db.commit()
        print("  -> Máquinas sincronizadas com sucesso.")

        # 2. Obter lista de painéis (do Excel ou JSON)
        products_data = []
        if os.path.exists(TEMPLATE_EXCEL_PATH):
            products_data = extract_products_from_excel(TEMPLATE_EXCEL_PATH)

        if not products_data and os.path.exists(JSON_CATALOG_PATH):
            print(f"Carregando fallback de: {JSON_CATALOG_PATH}")
            with open(JSON_CATALOG_PATH, "r", encoding="utf-8") as f:
                products_data = json.load(f)

        if products_data:
            # Salva cópia atualizada no JSON local
            with open(JSON_CATALOG_PATH, "w", encoding="utf-8") as f:
                json.dump(products_data, f, ensure_ascii=False, indent=2)
            print(f"  -> Catálogo espelhado em '{JSON_CATALOG_PATH}' ({len(products_data)} itens).")

            print(f"\n[2/2] Sincronizando Catálogo de Painéis no SQLite...")
            novos = 0
            atualizados = 0

            for p in products_data:
                code_int = int(p["code"])
                prod = db.query(models.Product).filter(models.Product.code == code_int).first()
                if not prod:
                    prod = models.Product(
                        code=code_int,
                        name=p["name"],
                        specification=p.get("specification") or p["name"],
                        dimensions=p.get("dimensions"),
                        unit_weight_kg=float(p.get("unit_weight_kg", 0.0)),
                        nominal_capacity_per_hour=float(p.get("nominal_capacity_per_hour", 0.0))
                    )
                    db.add(prod)
                    novos += 1
                else:
                    prod.name = p["name"]
                    prod.specification = p.get("specification") or p["name"]
                    prod.dimensions = p.get("dimensions")
                    prod.unit_weight_kg = float(p.get("unit_weight_kg", 0.0))
                    if p.get("nominal_capacity_per_hour"):
                        prod.nominal_capacity_per_hour = float(p["nominal_capacity_per_hour"])
                    atualizados += 1

            db.commit()
            print(f"  -> Catálogo no SQLite: {novos} novos | {atualizados} atualizados.")

        # Resumo final
        total_maq = db.query(models.Machine).count()
        total_prod = db.query(models.Product).count()
        total_entradas = db.query(models.ProductionEntry).count()

        print("\n" + "=" * 65)
        print("  STATUS ATUAL DO BANCO (SQLite):")
        print(f"  - Total de Máquinas: {total_maq}")
        print(f"  - Total de Produtos / Painéis no Catálogo: {total_prod}")
        print(f"  - Total de Apontamentos registrados: {total_entradas}")
        print("=" * 65)

    except Exception as e:
        db.rollback()
        print(f"\n[ERRO] Falha ao alimentar o banco: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()


if __name__ == "__main__":
    carregar_dados()

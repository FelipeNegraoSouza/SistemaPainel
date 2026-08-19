import os
import shutil
from datetime import datetime, time, timedelta
from typing import Dict, Any, Optional, List
import openpyxl
from sqlalchemy import or_, and_
from sqlalchemy.orm import Session

from backend import models

# Caminhos padrão do arquivo modelo e destino no drive Y:
DEFAULT_TEMPLATE_PATH = r"Y:\03 - PAINEL CORRUGADO\Ficha_apontamento_modelo_painel.xlsx"
DEFAULT_BASE_DEST_DIR = r"Y:\03 - PAINEL CORRUGADO\Fichas"

# Mapeamento dos meses para nomes de pasta em português
MONTH_NAMES = {
    1: "01 - JANEIRO",
    2: "02 - FEVEREIRO",
    3: "03 - MARÇO",
    4: "04 - ABRIL",
    5: "05 - MAIO",
    6: "06 - JUNHO",
    7: "07 - JULHO",
    8: "08 - AGOSTO",
    9: "09 - SETEMBRO",
    10: "10 - OUTUBRO",
    11: "11 - NOVEMBRO",
    12: "12 - DEZEMBRO"
}

MONTH_NAMES_LOWER = {
    1: "janeiro",
    2: "fevereiro",
    3: "março",
    4: "abril",
    5: "maio",
    6: "junho",
    7: "julho",
    8: "agosto",
    9: "setembro",
    10: "outubro",
    11: "novembro",
    12: "dezembro"
}

# Normalização de nomes de máquinas
MACHINE_NAME_MAP = {
    "Dobra 1": "DOBRA 01",
    "Dobra 2": "DOBRA 02",
    "Solda Lateral 1": "SOLDA LATERAL 01",
    "Solda Lateral 2": "SOLDA LATERAL 02",
    "Solda Ponto 1": "SOLDA PONTO 01",
    "Solda Ponto 2": "SOLDA PONTO 02",
    "Revisão": "REVISÃO",
    "Solda Manual": "SOLDA MANUAL"
}

def normalize_machine_name(name: str) -> str:
    if not name:
        return ""
    if name in MACHINE_NAME_MAP:
        return MACHINE_NAME_MAP[name]
    upper = name.strip().upper()
    if upper == "DOBRA 1": return "DOBRA 01"
    if upper == "DOBRA 2": return "DOBRA 02"
    if upper == "SOLDA LATERAL 1": return "SOLDA LATERAL 01"
    if upper == "SOLDA LATERAL 2": return "SOLDA LATERAL 02"
    if upper == "SOLDA PONTO 1": return "SOLDA PONTO 01"
    if upper == "SOLDA PONTO 2": return "SOLDA PONTO 02"
    return upper


def resolve_paths(reference_date: str) -> Dict[str, Any]:
    """
    Calcula os caminhos do arquivo modelo, pasta de destino e arquivo diário.
    """
    ref_dt = None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            ref_dt = datetime.strptime(reference_date.strip(), fmt)
            break
        except ValueError:
            continue

    if not ref_dt:
        ref_dt = datetime.now()

    year_str = str(ref_dt.year)
    month_folder = MONTH_NAMES.get(ref_dt.month, f"{ref_dt.month:02d}")
    
    file_name = ref_dt.strftime("%d-%m-%Y.xlsx")
    target_dir = os.path.join(DEFAULT_BASE_DEST_DIR, year_str, month_folder)
    target_filepath = os.path.join(target_dir, file_name)

    template_path = DEFAULT_TEMPLATE_PATH
    if not os.path.exists(template_path):
        alt_template = template_path.replace(".xlsx", "")
        if os.path.exists(alt_template):
            template_path = alt_template

    return {
        "date_obj": ref_dt,
        "date_iso": ref_dt.strftime("%Y-%m-%d"),
        "date_formatted": ref_dt.strftime("%d-%m-%Y"),
        "date_display": ref_dt.strftime("%d/%m/%Y"),
        "template_path": template_path,
        "template_exists": os.path.exists(template_path),
        "target_dir": target_dir,
        "target_filepath": target_filepath,
        "file_exists": os.path.exists(target_filepath),
        "file_name": file_name
    }


def ensure_daily_sheet_exists(reference_date: str) -> Dict[str, Any]:
    """
    Copia o modelo limpo para a pasta de destino caso o arquivo do dia não exista.
    """
    paths = resolve_paths(reference_date)
    target_filepath = paths["target_filepath"]
    target_dir = paths["target_dir"]
    template_path = paths["template_path"]

    if not os.path.exists(target_filepath):
        if not os.path.exists(template_path):
            raise FileNotFoundError(
                f"Arquivo modelo não encontrado em '{template_path}'. Verifique o acesso ao drive Y:."
            )
        
        os.makedirs(target_dir, exist_ok=True)
        shutil.copy2(template_path, target_filepath)
        paths["file_exists"] = True
        paths["created_now"] = True
    else:
        paths["created_now"] = False

    return paths


# Cache em memória do catálogo da aba BD para máxima performance
_CACHED_PRODUCT_LOOKUP: Optional[Dict[str, Dict[str, float]]] = None

def get_product_lookup(template_path: str) -> Dict[str, Dict[str, float]]:
    global _CACHED_PRODUCT_LOOKUP
    if _CACHED_PRODUCT_LOOKUP is not None:
        return _CACHED_PRODUCT_LOOKUP

    lookup = {}
    if not os.path.exists(template_path):
        return lookup

    try:
        wb_bd = openpyxl.load_workbook(template_path, data_only=True)
        if "BD" in wb_bd.sheetnames:
            ws_bd = wb_bd["BD"]
            for r in range(2, ws_bd.max_row + 1):
                code = ws_bd.cell(r, 3).value
                desc = str(ws_bd.cell(r, 4).value or "").strip()
                medida = str(ws_bd.cell(r, 6).value or "").strip()
                peso = ws_bd.cell(r, 7).value or 0.0
                dobra = ws_bd.cell(r, 10).value or 0.0
                solda_l = ws_bd.cell(r, 11).value or 0.0
                solda_p = ws_bd.cell(r, 12).value or 0.0
                
                info = {
                    "peso": float(peso) if isinstance(peso, (int, float)) else 0.0,
                    "dobra": float(dobra) if isinstance(dobra, (int, float)) else 0.0,
                    "solda_l": float(solda_l) if isinstance(solda_l, (int, float)) else 0.0,
                    "solda_p": float(solda_p) if isinstance(solda_p, (int, float)) else 0.0,
                }
                if code:
                    lookup[str(code).strip()] = info
                if desc:
                    lookup[desc.upper()] = info
                if medida and medida != "#VALUE!":
                    lookup[medida.upper()] = info
        wb_bd.close()
    except Exception as e:
        print(f"[Aviso Excel] Erro ao construir lookup de produtos: {e}")

    _CACHED_PRODUCT_LOOKUP = lookup
    return _CACHED_PRODUCT_LOOKUP


def sync_date_to_excel(reference_date: str, db: Session) -> Dict[str, Any]:
    """
    Processa e calcula TODOS os valores no backend Python e grava os resultados calculados
    diretamente na planilha diária do Excel (aba BD_LANÇAMENTOS) sem uso de fórmulas de modelo.
    O Excel passa a atuar estritamente como visualizador e relatório de impressão.
    """
    paths = ensure_daily_sheet_exists(reference_date)
    target_filepath = paths["target_filepath"]
    ref_date_iso = paths["date_iso"]
    date_display = paths["date_display"]
    date_obj = paths["date_obj"]
    mes_nome = MONTH_NAMES_LOWER.get(date_obj.month, "")

    # Calcula data útil anterior para o turno noturno (Segunda -> Sexta)
    weekday = date_obj.weekday()
    if weekday == 0:
        prev_date_obj = date_obj - timedelta(days=3)
    elif weekday == 6:
        prev_date_obj = date_obj - timedelta(days=2)
    else:
        prev_date_obj = date_obj - timedelta(days=1)
    prev_date_iso = prev_date_obj.strftime("%Y-%m-%d")

    # Consulta sessões e lançamentos da planilha diária no SQLite:
    # 1. Turno Diurno trabalhado na data
    # 2. Turno Noturno trabalhado na noite do dia útil anterior
    sessions = db.query(models.ProductionSession).filter(
        or_(
            and_(models.ProductionSession.reference_date == ref_date_iso, models.ProductionSession.shift == "Diurno"),
            and_(models.ProductionSession.reference_date == prev_date_iso, models.ProductionSession.shift == "Noturno"),
            and_(models.ProductionSession.reference_date == ref_date_iso, models.ProductionSession.shift == "Noturno")
        )
    ).all()

    # Carrega a planilha Excel diária
    wb = openpyxl.load_workbook(target_filepath)
    bd_lookup = get_product_lookup(paths["template_path"])
    
    # Encontra a aba BD_LANÇAMENTOS
    ws = None
    for sheet in wb.worksheets:
        if "BD_LAN" in sheet.title.upper() or "LANCAMENTOS" in sheet.title.upper():
            ws = sheet
            break
    if ws is None:
        ws = wb.worksheets[4] if len(wb.worksheets) >= 5 else wb.active

    # Atualiza cabeçalho do topo (B1)
    cell_b1 = ws.cell(1, 2)
    if not isinstance(cell_b1, openpyxl.cell.cell.MergedCell):
        cell_b1.value = f"PAINEL - {date_display}"

    # Limpa linhas anteriores dos lançamentos (linhas 3 a 28, colunas B a S)
    for r in range(3, 29):
        for c in range(2, 20):
            cell = ws.cell(r, c)
            if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                cell.value = None

    # Coleta todas as entradas para processamento
    all_entries = []
    for s in sessions:
        machine_name = s.machine.name if s.machine else f"Máquina {s.machine_id}"
        norm_machine = normalize_machine_name(machine_name)
        operator_name = s.operator_name or ""
        
        for e in s.entries:
            all_entries.append({
                "machine": norm_machine,
                "product_spec": e.product_spec_custom or (e.product.name if e.product else ""),
                "product_code": e.product_code,
                "product_obj": e.product,
                "operator": operator_name,
                "qty": e.qty_produced or 0,
                "start_time": e.start_time,
                "end_time": e.end_time,
                "gross_minutes": e.gross_minutes or 0,
                "stop_minutes": e.total_stop_minutes or 0,
                "net_minutes": e.net_minutes or 0
            })

    # Ordena por máquina e hora de início
    all_entries.sort(key=lambda x: (x["machine"], x["start_time"]))

    # Totalizadores para os cartões de KPI (linhas 29 a 45)
    total_peso_dobra = 0.0
    total_peso_solda_l = 0.0
    total_peso_solda_p = 0.0
    total_peso_outros = 0.0

    row_idx = 3
    for entry in all_entries:
        if row_idx > 28:
            break

        machine = entry["machine"]
        product_spec = entry["product_spec"].strip()
        operator = entry["operator"]
        qty = entry["qty"]
        product_code = entry["product_code"]
        product_obj = entry["product_obj"]

        # 1. Obter Peso Unitário e Capacidade Nominal do Produto via Backend
        peso_unitario = 0.0
        dobra_cap = 0.0
        solda_l_cap = 0.0
        solda_p_cap = 0.0

        # Tenta pelo objeto do banco
        if product_obj and product_obj.unit_weight_kg:
            peso_unitario = float(product_obj.unit_weight_kg)
            dobra_cap = float(product_obj.nominal_capacity_per_hour or 0.0)
            solda_l_cap = float(product_obj.nominal_capacity_per_hour or 0.0)
            solda_p_cap = float(product_obj.nominal_capacity_per_hour or 0.0)

        # Fallback na tabela BD do Excel
        if peso_unitario == 0.0:
            info = None
            if product_code and str(product_code) in bd_lookup:
                info = bd_lookup[str(product_code)]
            elif product_spec.upper() in bd_lookup:
                info = bd_lookup[product_spec.upper()]
            else:
                for k, v in bd_lookup.items():
                    if k in product_spec.upper() or product_spec.upper() in k:
                        info = v
                        break
            if info:
                peso_unitario = info["peso"]
                if dobra_cap == 0.0: dobra_cap = info["dobra"]
                if solda_l_cap == 0.0: solda_l_cap = info["solda_l"]
                if solda_p_cap == 0.0: solda_p_cap = info["solda_p"]

        # Cálculo do Peso Total
        peso_total = round(peso_unitario * qty, 2) if peso_unitario > 0 else 0.0

        # 2. Horários e Cálculo de Tempo
        try:
            start_parts = [int(p) for p in entry["start_time"].split(":")]
            end_parts = [int(p) for p in entry["end_time"].split(":")]
            start_dt = date_obj.replace(hour=start_parts[0], minute=start_parts[1], second=0)
            end_dt = date_obj.replace(hour=end_parts[0], minute=end_parts[1], second=0)
            if end_dt <= start_dt:
                end_dt += timedelta(days=1)
            gross_minutes = int((end_dt - start_dt).total_seconds() / 60)
        except Exception:
            start_dt = entry["start_time"]
            end_dt = entry["end_time"]
            gross_minutes = entry["gross_minutes"] or 0

        stop_mins = entry["stop_minutes"]
        net_minutes = max(0, gross_minutes - stop_mins)
        net_hours_decimal = round(net_minutes / 60.0, 2)

        # 3. Operação, Agrupamento e Capacidade por Máquina
        operacao = ""
        coluna1 = ""
        nominal_rate_per_hour = 0.0

        if "DOBRA" in machine:
            operacao = "DOBRA"
            coluna1 = "DOBRA"
            nominal_rate_per_hour = dobra_cap
            total_peso_dobra += peso_total
        elif "SOLDA LATERAL" in machine:
            operacao = "SOLDA LATERAL"
            coluna1 = "SOLDA LATERAL"
            nominal_rate_per_hour = solda_l_cap
            total_peso_solda_l += peso_total
        elif "SOLDA PONTO" in machine:
            operacao = "SOLDA PONTO"
            coluna1 = "SOLDA PONTO"
            nominal_rate_per_hour = solda_p_cap
            total_peso_solda_p += peso_total
        elif "REVISÃO" in machine:
            operacao = "REVISÃO"
            coluna1 = "REVISÃO"
            total_peso_outros += peso_total
        elif "SOLDA MANUAL" in machine:
            operacao = "SOLDA MANUAL"
            coluna1 = "SOLDA MANUAL"
            total_peso_outros += peso_total
        else:
            operacao = machine.split()[0] if machine else ""
            coluna1 = operacao
            total_peso_outros += peso_total

        # 4. Cálculo da Capacidade Nominal e Produtividade (%)
        cap_producao = round(nominal_rate_per_hour * (net_minutes / 60.0), 2)
        produtividade_pct = round(qty / cap_producao, 4) if cap_producao > 0 else None

        # --- GRAVAÇÃO DOS VALORES CALCULADOS DIRETAMENTE NAS CÉLULAS (SEM FÓRMULAS) ---
        
        # B: MAQUINA
        ws.cell(row_idx, 2).value = machine
        # C: PRODUTO
        ws.cell(row_idx, 3).value = product_spec
        # D: OPERADOR
        ws.cell(row_idx, 4).value = operator
        # E: QTD PRODUZIDA
        ws.cell(row_idx, 5).value = qty
        # F: PESO CALCULADO
        ws.cell(row_idx, 6).value = peso_total if peso_total > 0 else (0.0 if qty > 0 else None)
        # G: HORA INÍCIO
        ws.cell(row_idx, 7).value = start_dt
        # H: HORA FINAL
        ws.cell(row_idx, 8).value = end_dt
        # I: parada
        if stop_mins > 0:
            ws.cell(row_idx, 9).value = time(hour=stop_mins // 60, minute=stop_mins % 60)
        else:
            ws.cell(row_idx, 9).value = None
        # J: HORAS TRABALHADAS (Tempo líquido)
        ws.cell(row_idx, 10).value = time(hour=net_minutes // 60, minute=net_minutes % 60)
        # K: CAP. PRODUÇÃO
        ws.cell(row_idx, 11).value = cap_producao if cap_producao > 0 else (0 if nominal_rate_per_hour == 0 else None)
        # L: PRODUT. %
        ws.cell(row_idx, 12).value = produtividade_pct
        # M: TOTAL HORA (Minutos líquidos)
        ws.cell(row_idx, 13).value = net_minutes
        # N: HORA/DIA (Horas líquidas decimais)
        ws.cell(row_idx, 14).value = net_hours_decimal
        # O: MÊS
        ws.cell(row_idx, 15).value = mes_nome
        # P: DATA
        ws.cell(row_idx, 16).value = date_obj.date()
        # Q: OPERAÇÃO
        ws.cell(row_idx, 17).value = operacao
        # R: Coluna1
        ws.cell(row_idx, 18).value = coluna1
        # S: Coluna2 (Capacidade por hora da máquina)
        ws.cell(row_idx, 19).value = nominal_rate_per_hour

        row_idx += 1

    # --- ATUALIZAÇÃO DOS CARTÕES DE KPI / TOTAIS NO FINAL DA PLANILHA ---
    
    # Realizado do dia (Solda Lateral ou Total de Produção)
    realizado_dia = round(total_peso_solda_l if total_peso_solda_l > 0 else (total_peso_dobra + total_peso_solda_p), 2)
    
    # Meta Dia (Linha 31, Coluna B)
    meta_dia_cell = ws.cell(31, 2)
    meta_dia = float(meta_dia_cell.value) if meta_dia_cell.value and isinstance(meta_dia_cell.value, (int, float)) else 2500.0
    
    # Realizado (Linha 31, Coluna G)
    ws.cell(31, 7).value = realizado_dia
    # % Meta Dia (Linha 31, Coluna J)
    ws.cell(31, 10).value = round(realizado_dia / meta_dia, 4) if meta_dia > 0 else 0.0

    # Totais por Operação (Linhas 43 a 45)
    ws.cell(43, 8).value = round(total_peso_dobra, 2)
    ws.cell(44, 8).value = round(total_peso_solda_l, 2)
    ws.cell(45, 8).value = round(total_peso_solda_p, 2)

    # Ajusta o range da Tabela5 para manter formatação de impressão
    if ws.tables:
        for tbl in ws.tables.values():
            if tbl.name == "Tabela5":
                tbl.ref = f"B2:T28"

    wb.save(target_filepath)

    return {
        "success": True,
        "synced_entries_count": len(all_entries),
        "target_filepath": target_filepath,
        "date_formatted": paths["date_formatted"],
        "date_iso": ref_date_iso,
        "processed_metrics": {
            "realizado_dia_kg": realizado_dia,
            "dobra_kg": round(total_peso_dobra, 2),
            "solda_lateral_kg": round(total_peso_solda_l, 2),
            "solda_ponto_kg": round(total_peso_solda_p, 2)
        }
    }


def sync_catalog_from_excel_bd(db: Session, template_path: Optional[str] = None) -> Dict[str, Any]:
    """
    Lê a aba 'BD' da planilha modelo no drive Y: e sincroniza com a tabela 'products' do SQLite.
    Salva também uma cópia no JSON local de catálogo.
    """
    import re
    import json

    t_path = template_path or DEFAULT_TEMPLATE_PATH
    if not os.path.exists(t_path):
        alt_path = t_path.replace(".xlsx", "")
        if os.path.exists(alt_path):
            t_path = alt_path
        else:
            raise FileNotFoundError(f"Arquivo modelo não encontrado em: {t_path}")

    wb = openpyxl.load_workbook(t_path, data_only=True)
    if "BD" not in wb.sheetnames:
        raise ValueError("Aba 'BD' não encontrada na planilha modelo.")

    ws = wb["BD"]
    rows = list(ws.iter_rows(values_only=True))[1:]

    synced_products = []
    seen_codes = set()
    new_count = 0
    updated_count = 0

    for r in rows:
        raw_code = r[2]      # Coluna C
        raw_desc = r[3]      # Coluna D
        raw_base = r[5]      # Coluna F
        raw_weight = r[6]    # Coluna G

        if raw_code is None or str(raw_code).strip() in ("", "#N/A", "0", "None"):
            continue

        try:
            code_int = int(float(str(raw_code).replace(",", ".")))
        except (ValueError, TypeError):
            continue

        if code_int <= 0 or code_int in seen_codes:
            continue

        seen_codes.add(code_int)
        desc = str(raw_desc).strip() if raw_desc is not None else ""
        if not desc or desc in ("#N/A", "None"):
            continue

        # Medida Base (Coluna F)
        base_dim = str(raw_base).strip() if raw_base is not None else ""
        if "#VALUE" in base_dim or "#N/A" in base_dim or not base_dim or base_dim == "None":
            m = re.search(r"(\d+\s*[xX]\s*\d+\s*[xX]\s*\d+)", desc)
            if m:
                base_dim = m.group(1).upper().strip()
            else:
                base_dim = desc.replace("P. CORRUGADO", "").replace("P.CORRUGADO", "").strip()

        # Peso (Coluna G)
        try:
            weight = float(str(raw_weight).replace(",", ".")) if raw_weight is not None else 0.0
            if weight < 0:
                weight = 0.0
        except (ValueError, TypeError):
            weight = 0.0

        item_dict = {
            "code": code_int,
            "name": desc,
            "specification": desc,
            "dimensions": base_dim,
            "unit_weight_kg": round(weight, 2),
            "nominal_capacity_per_hour": 0.0
        }
        synced_products.append(item_dict)

        # Atualiza ou insere no banco
        prod = db.query(models.Product).filter(models.Product.code == code_int).first()
        if not prod:
            prod = models.Product(
                code=code_int,
                name=desc,
                specification=desc,
                dimensions=base_dim,
                unit_weight_kg=round(weight, 2),
                nominal_capacity_per_hour=0.0
            )
            db.add(prod)
            new_count += 1
        else:
            prod.name = desc
            prod.specification = desc
            prod.dimensions = base_dim
            prod.unit_weight_kg = round(weight, 2)
            updated_count += 1

    db.commit()

    # Atualiza JSON local de backup
    json_path = os.path.join(os.path.dirname(__file__), "..", "scripts", "catalogo_paineis.json")
    try:
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(synced_products, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[Aviso] Falha ao salvar backup JSON do catálogo: {e}")

    return {
        "success": True,
        "total_extracted": len(synced_products),
        "new_count": new_count,
        "updated_count": updated_count,
        "template_path": t_path
    }

